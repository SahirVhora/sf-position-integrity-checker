"""
remediation.py - Auto-Remediation Payload Generator

For each position integrity issue, derive the correct field value from the
existing foundation lookups and build an OData v2 Position PATCH payload.

Dry-run (default): writes payloads to output/ as JSON + Excel. No OData calls.
Live apply (dry_run=False): POSTs each payload to the SF tenant. Requires
explicit confirmation from the caller.
"""

import datetime
import json
import os
from dataclasses import asdict, dataclass, field
from typing import Any

# ---------------------------------------------------------------------------
# Per-check remediation rules
# Defines how to derive the correct value for each check ID.
#
# Each entry: check_id -> {
#   "position_field":   position field to update,
#   "source_lookup":    lookups key for the foundation record,
#   "source_code_field":position field used as key into source_lookup,
#   "source_value_field":field on the foundation record with the correct value,
#   "junction_source":  if set, correct value comes from a junction set (set_membership checks)
# }
# ---------------------------------------------------------------------------

_REMEDIATION_RULES: dict[str, dict[str, str]] = {
    # CHK-01: Sub Dept's cust_Department must match Position's department
    "CHK-01": {
        "position_field": "department",
        "source_lookup": "sub_departments",
        "source_code_field": "cust_subDepartment",
        "source_value_field": "cust_Department",
    },
    # CHK-02: Department's cust_Division must match Position's division
    "CHK-02": {
        "position_field": "division",
        "source_lookup": "departments",
        "source_code_field": "department",
        "source_value_field": "cust_Division",
    },
    # CHK-03: Division linked BUs must include Position's businessUnit
    # Remediation: update position.businessUnit to first valid linked BU
    "CHK-03": {
        "position_field": "businessUnit",
        "junction_source": "div_to_bus",
        "source_code_field": "division",
    },
    # CHK-04: BU linked LEs must include Position's company
    # Remediation: update position.company to first valid linked LE
    "CHK-04": {
        "position_field": "company",
        "junction_source": "bu_to_les",
        "source_code_field": "businessUnit",
    },
    # CHK-05: Cost Centre linked BUs must include Position's businessUnit
    # Remediation: update position.businessUnit to first valid linked BU
    "CHK-05": {
        "position_field": "businessUnit",
        "junction_source": "cc_to_bus",
        "source_code_field": "costCenter",
    },
    # CHK-06: Job Code jobFunction must match Position's cust_JobFunction
    "CHK-06": {
        "position_field": "cust_JobFunction",
        "source_lookup": "job_codes",
        "source_code_field": "jobCode",
        "source_value_field": "jobFunction",
    },
    # CHK-07: Job Code sub-function must match Position's cust_jobSubFunction
    "CHK-07": {
        "position_field": "cust_jobSubFunction",
        "source_lookup": "job_codes",
        "source_code_field": "jobCode",
        "source_value_field": "cust_jobsubfunction",
    },
    # CHK-08: Job Code grade must match Position's cust_GlobalJobLevel
    "CHK-08": {
        "position_field": "cust_GlobalJobLevel",
        "source_lookup": "job_codes",
        "source_code_field": "jobCode",
        "source_value_field": "grade",
    },
    # CHK-09: Job Code career path must match Position's cust_CareerPath
    "CHK-09": {
        "position_field": "cust_CareerPath",
        "source_lookup": "job_codes",
        "source_code_field": "jobCode",
        "source_value_field": "cust_careerPath",
    },
}


@dataclass
class RemedEntry:
    position_code: str
    position_title: str
    check_id: str
    position_field: str
    old_value: str
    new_value: str
    confidence: str  # "HIGH" or "MEDIUM"
    payload: dict = field(default_factory=dict)
    skipped: bool = False
    skip_reason: str = ""


@dataclass
class ApplyResult:
    total: int = 0
    applied: int = 0
    skipped: int = 0
    failed: int = 0
    dry_run: bool = True
    entries: list[RemedEntry] = field(default_factory=list)
    apply_log: list[dict] = field(default_factory=list)


def build_remediation_pack(
    issues: list[dict[str, Any]],
    lookups: dict[str, Any],
    positions: list[dict[str, Any]],
) -> list[RemedEntry]:
    """
    For each issue with a known remediation rule, derive the correct value
    from lookups and build an OData v2 Position PATCH payload.

    Issues for CHK-10 to CHK-17 (foundation_active checks) are skipped -
    those require fixing the foundation object, not the position.

    Returns a list of RemedEntry objects (one per actionable issue).
    """
    pos_map: dict[str, dict] = {p["code"]: p for p in positions if p.get("code")}

    entries: list[RemedEntry] = []

    for issue in issues:
        check_id = issue.get("Check ID", "")
        rule = _REMEDIATION_RULES.get(check_id)

        if rule is None:
            # CHK-10 to CHK-17 (foundation_active) - can't fix position for these
            continue

        position_code = issue.get("Position ID", "")
        position_title = issue.get("Position Title", "")
        pos = pos_map.get(position_code)

        if not pos:
            entries.append(
                RemedEntry(
                    position_code=position_code,
                    position_title=position_title,
                    check_id=check_id,
                    position_field=rule["position_field"],
                    old_value=issue.get("Issue Description", ""),
                    new_value="",
                    confidence="",
                    skipped=True,
                    skip_reason="Position not found in current extract",
                )
            )
            continue

        old_value = pos.get(rule["position_field"], "") or ""
        entry = _build_entry(issue, pos, rule, lookups, position_title, old_value)
        entries.append(entry)

    return entries


def _build_entry(
    issue: dict,
    pos: dict,
    rule: dict,
    lookups: dict,
    position_title: str,
    old_value: str,
) -> RemedEntry:
    """Derive the correct value and build one RemedEntry."""
    check_id = issue["Check ID"]
    position_field = rule["position_field"]
    source_code_field = rule["source_code_field"]
    source_code = pos.get(source_code_field, "")

    if not source_code:
        return RemedEntry(
            position_code=pos["code"],
            position_title=position_title,
            check_id=check_id,
            position_field=position_field,
            old_value=old_value,
            new_value="",
            confidence="",
            skipped=True,
            skip_reason=f"Position has no value for {source_code_field}",
        )

    if "junction_source" in rule:
        # set_membership check: correct value is from the junction set
        junction_key = rule["junction_source"]
        valid_set = lookups.get(junction_key, {}).get(source_code, set())
        if not valid_set:
            return RemedEntry(
                position_code=pos["code"],
                position_title=position_title,
                check_id=check_id,
                position_field=position_field,
                old_value=old_value,
                new_value="",
                confidence="",
                skipped=True,
                skip_reason=f"No valid linked codes found in {junction_key} for {source_code!r}",
            )
        # Sort for determinism; MEDIUM if multiple options
        sorted_options = sorted(valid_set)
        correct_value = sorted_options[0]
        confidence = "HIGH" if len(sorted_options) == 1 else "MEDIUM"
    else:
        # scalar_match check: correct value is a field on the foundation record
        lookup_key = rule["source_lookup"]
        source_value_field = rule["source_value_field"]
        fo_record = lookups.get(lookup_key, {}).get(source_code)
        if not fo_record:
            return RemedEntry(
                position_code=pos["code"],
                position_title=position_title,
                check_id=check_id,
                position_field=position_field,
                old_value=old_value,
                new_value="",
                confidence="",
                skipped=True,
                skip_reason=f"Foundation record {source_code!r} not in {lookup_key} lookup",
            )
        correct_value = fo_record.get(source_value_field, "") or ""
        if not correct_value:
            return RemedEntry(
                position_code=pos["code"],
                position_title=position_title,
                check_id=check_id,
                position_field=position_field,
                old_value=old_value,
                new_value="",
                confidence="",
                skipped=True,
                skip_reason=f"{source_value_field} is blank on foundation record {source_code!r}",
            )
        confidence = "HIGH"

    payload = _build_payload(pos, position_field, correct_value)

    return RemedEntry(
        position_code=pos["code"],
        position_title=position_title,
        check_id=check_id,
        position_field=position_field,
        old_value=old_value,
        new_value=correct_value,
        confidence=confidence,
        payload=payload,
    )


def _build_payload(pos: dict, field_to_fix: str, correct_value: str) -> dict:
    """
    Build an OData v2 Position PATCH payload.
    Includes the effective date as required by SF effective-dated entities.
    """
    effective_start = pos.get("effectiveStartDate", "") or ""
    # Convert stored YYYY-MM-DD to SF /Date(ms)/ format if needed
    start_date_ms = _date_to_epoch_ms(effective_start) if effective_start else None

    payload: dict[str, Any] = {
        "code": pos["code"],
    }
    if start_date_ms is not None:
        payload["effectiveStartDate"] = f"/Date({start_date_ms})/"

    payload[field_to_fix] = correct_value
    return payload


def _date_to_epoch_ms(date_str: str) -> int | None:
    """Convert YYYY-MM-DD string to epoch milliseconds."""
    try:
        d = datetime.date.fromisoformat(date_str[:10])
        epoch = datetime.date(1970, 1, 1)
        return int((d - epoch).total_seconds() * 1000)
    except (ValueError, TypeError):
        return None


def apply_remediation(
    entries: list[RemedEntry],
    country: str = "CA",
    dry_run: bool = True,
) -> ApplyResult:
    """
    Write remediation pack (JSON + Excel) and optionally POST to SF.

    dry_run=True  (default): writes output files only. No OData calls.
    dry_run=False:  POSTs each non-skipped payload to Position entity.
                    Requires explicit caller confirmation (pass dry_run=False).
    """
    os.makedirs("output", exist_ok=True)
    today = datetime.date.today().strftime("%Y%m%d")
    json_path = os.path.join("output", f"remediation_pack_{country}_{today}.json")
    xlsx_path = os.path.join("output", f"remediation_pack_{country}_{today}.xlsx")
    audit_path = os.path.join("output", f"remediation_audit_{country}_{today}.jsonl")

    result = ApplyResult(
        total=len(entries),
        dry_run=dry_run,
        entries=entries,
    )

    actionable = [e for e in entries if not e.skipped]
    result.skipped = len(entries) - len(actionable)

    # Always write JSON pack
    _write_json_pack(entries, json_path, dry_run, country)

    # Always write Excel pack
    _write_excel_pack(entries, xlsx_path, country)

    if not dry_run:
        result = _post_payloads(result, actionable, audit_path, country)
    else:
        result.applied = 0
        result.failed = 0

    return result


def _write_json_pack(
    entries: list[RemedEntry],
    path: str,
    dry_run: bool,
    country: str,
) -> None:
    doc = {
        "schema": "sf-remediation-pack/v1",
        "generated_at": datetime.datetime.now(datetime.UTC).isoformat(),
        "country": country,
        "dry_run": dry_run,
        "total": len(entries),
        "skipped": sum(1 for e in entries if e.skipped),
        "actionable": sum(1 for e in entries if not e.skipped),
        "entries": [asdict(e) for e in entries],
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, indent=2, default=str)


def _write_excel_pack(entries: list[RemedEntry], path: str, country: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return  # openpyxl optional for this module

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws_sum = wb.active
    ws_sum.title = "Remediation Summary"
    header_fill = PatternFill("solid", fgColor="1B2A4A")
    header_font = Font(color="FFFFFF", bold=True)

    total = len(entries)
    skipped = sum(1 for e in entries if e.skipped)
    actionable = total - skipped
    high = sum(1 for e in entries if not e.skipped and e.confidence == "HIGH")
    medium = sum(1 for e in entries if not e.skipped and e.confidence == "MEDIUM")

    check_counts: dict[str, int] = {}
    for e in entries:
        if not e.skipped:
            check_counts[e.check_id] = check_counts.get(e.check_id, 0) + 1

    ws_sum.append(["SF Position Integrity Checker - Remediation Pack"])
    ws_sum.append(["Country", country])
    ws_sum.append(["Generated", datetime.datetime.now(datetime.UTC).strftime("%Y-%m-%d %H:%M UTC")])
    ws_sum.append([])
    ws_sum.append(["Total Issues", total])
    ws_sum.append(["Actionable", actionable])
    ws_sum.append(["Skipped", skipped])
    ws_sum.append(["HIGH Confidence", high])
    ws_sum.append(["MEDIUM Confidence", medium])
    ws_sum.append([])
    ws_sum.append(["Check ID", "Actionable Count"])
    for cid, cnt in sorted(check_counts.items()):
        ws_sum.append([cid, cnt])

    # --- Sheet 2: Payloads ---
    ws_pay = wb.create_sheet("Payloads")
    cols = [
        "Position Code",
        "Position Title",
        "Check ID",
        "Position Field",
        "Old Value",
        "New Value",
        "Confidence",
        "Skipped",
        "Skip Reason",
        "Payload JSON",
    ]
    ws_pay.append(cols)
    for cell in ws_pay[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    high_fill = PatternFill("solid", fgColor="C6EFCE")
    med_fill = PatternFill("solid", fgColor="FFEB9C")
    skip_fill = PatternFill("solid", fgColor="F2F2F2")

    for e in entries:
        row = [
            e.position_code,
            e.position_title,
            e.check_id,
            e.position_field,
            e.old_value,
            e.new_value,
            e.confidence,
            "Yes" if e.skipped else "No",
            e.skip_reason,
            json.dumps(e.payload) if e.payload else "",
        ]
        ws_pay.append(row)
        last_row = ws_pay.max_row
        fill = skip_fill if e.skipped else (high_fill if e.confidence == "HIGH" else med_fill)
        for cell in ws_pay[last_row]:
            cell.fill = fill

    # --- Sheet 3: Apply Log (placeholder) ---
    ws_log = wb.create_sheet("Apply Log")
    ws_log.append(["Timestamp", "Position Code", "Check ID", "HTTP Status", "Result", "Detail"])
    ws_log["A1"].comment = None  # populated by _post_payloads if dry_run=False

    wb.save(path)


def _post_payloads(
    result: ApplyResult,
    actionable: list[RemedEntry],
    audit_path: str,
    country: str,
) -> ApplyResult:
    """POST payloads to the SF tenant. Appends to JSONL audit log."""
    import database
    from api_client import fetch_all  # noqa: F401 - import to validate availability

    database.set_country(country)

    try:
        from api_client import _get_client  # internal; may not exist

        client = _get_client()
    except (ImportError, AttributeError):
        # Fall back to direct requests if _get_client not exposed
        from api_client import _build_session  # type: ignore

        client = _build_session()

    with open(audit_path, "a", encoding="utf-8") as log_fh:
        for entry in actionable:
            ts = datetime.datetime.now(datetime.UTC).isoformat()
            try:
                resp = client.patch(
                    f"Position('{entry.position_code}')",
                    json=entry.payload,
                )
                http_status = getattr(resp, "status_code", "unknown")
                success = str(http_status).startswith("2")
                log_line = {
                    "timestamp": ts,
                    "position_code": entry.position_code,
                    "check_id": entry.check_id,
                    "field": entry.position_field,
                    "old_value": entry.old_value,
                    "new_value": entry.new_value,
                    "http_status": http_status,
                    "result": "OK" if success else "FAILED",
                }
                log_fh.write(json.dumps(log_line) + "\n")
                result.apply_log.append(log_line)
                if success:
                    result.applied += 1
                else:
                    result.failed += 1
            except Exception as exc:
                log_line = {
                    "timestamp": ts,
                    "position_code": entry.position_code,
                    "check_id": entry.check_id,
                    "field": entry.position_field,
                    "result": "ERROR",
                    "detail": str(exc),
                }
                log_fh.write(json.dumps(log_line) + "\n")
                result.apply_log.append(log_line)
                result.failed += 1

    return result
