"""
reporters.py - Generate Excel, CSV, HTML reports and a run manifest.
"""

import datetime
import json
import os
import re
from collections import Counter
from typing import Any, Dict, List

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
VERSION = "1.0.0"
GITHUB_URL = "github.com/sahirvhora/sf-position-integrity-checker"

COLUMNS = [
    "Position ID",
    "Position Title",
    "Effective Start Date",
    "Legal Entity",
    "Business Unit",
    "Division",
    "Department",
    "Sub Department",
    "Job Code",
    "Cost Centre",
    "Location",
    "Employee ID",
    "Employee Status",
    "Check ID",
    "Check Category",
    "Failed Field",
    "Issue Description",
    "Severity",
]

# Colours
HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # navy blue
CRITICAL_FILL = PatternFill("solid", fgColor="FFD7D7")  # light red
HIGH_FILL = PatternFill("solid", fgColor="FFE4C4")  # light orange
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="DCE6F1")


def _ensure_output_dir() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _datestamp() -> str:
    return datetime.date.today().strftime("%Y%m%d")


def _fmt_date(raw: Any) -> str:
    """Convert any SF date value to MM/DD/YYYY string. Returns original string on failure."""
    if not raw:
        return ""
    s = str(raw).strip()
    if s.startswith("/Date("):
        ms_part = s[6:].split(")")[0].split("+")[0].split("-")[0]
        try:
            d = datetime.date.fromtimestamp(int(ms_part) / 1000)
            return d.strftime("%m/%d/%Y")
        except (ValueError, OSError):
            return s
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            return datetime.datetime.strptime(s[: len(fmt)], fmt).strftime("%m/%d/%Y")
        except ValueError:
            pass
    try:
        return datetime.date.fromisoformat(s[:10]).strftime("%m/%d/%Y")
    except ValueError:
        return s


def _normalise_dates(issues: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Return a copy of issues with Effective Start Date formatted as MM/DD/YYYY."""
    out = []
    for row in issues:
        r = dict(row)
        r["Effective Start Date"] = _fmt_date(r.get("Effective Start Date", ""))
        out.append(r)
    return out


def _visible_issues(issues: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Return only issues whose rule has visible: true (default)."""
    from validators import CHECK_META

    return [
        i
        for i in issues
        if CHECK_META.get(i.get("Check ID", ""), {}).get("visible", True)
    ]


def _df_from_issues(issues: List[Dict[str, Any]]) -> pd.DataFrame:
    normalised = _normalise_dates(issues)
    if normalised:
        return pd.DataFrame(normalised, columns=COLUMNS)
    return pd.DataFrame(columns=COLUMNS)


def _mask_tenant_url(url: str) -> str:
    """Replace tenant subdomain with ***masked*** for safe logging."""
    return re.sub(r"(https?://)([^.]+)(\..*)", r"\1***masked***\3", url)


def _instance_name(tenant_url: str, instance_id: str = "") -> str:
    """Return the SF instance identifier.
    Uses instance_id (Company ID from config) when provided; falls back to
    parsing the hostname from tenant_url for backwards compatibility."""
    if instance_id:
        return instance_id
    if not tenant_url:
        return ""
    m = re.match(r"https?://([^./]+)", tenant_url.strip())
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def write_findings_json(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str = "CA",
    tenant_url: str = "",
    as_of_date: datetime.date | None = None,
) -> str:
    """Write findings in the SF Compass suite schema (sf-compass-findings/v1)."""
    _ensure_output_dir()
    path = os.path.join(
        OUTPUT_DIR, f"position_integrity_findings_{country}_{_datestamp()}.json"
    )
    normalised = _normalise_dates(issues)
    findings = []
    by_severity: Dict[str, int] = {}
    for issue in normalised:
        severity = str(issue.get("Severity", "")).strip().lower() or "info"
        by_severity[severity] = by_severity.get(severity, 0) + 1
        findings.append(
            {
                "id": issue.get("Check ID", ""),
                "severity": severity,
                "category": issue.get("Check Category", ""),
                "object_type": "Position",
                "object_id": issue.get("Position ID", ""),
                "field": issue.get("Failed Field", ""),
                "message": issue.get("Issue Description", ""),
                "details": {
                    "legal_entity": issue.get("Legal Entity", ""),
                    "department": issue.get("Department", ""),
                    "job_code": issue.get("Job Code", ""),
                    "effective_start_date": issue.get("Effective Start Date", ""),
                },
            }
        )
    document = {
        "schema": "sf-compass-findings/v1",
        "tool": "sf-position-integrity-checker",
        "tool_version": VERSION,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "tenant": _mask_tenant_url(tenant_url) if tenant_url else "",
        "scope": {
            "country": country,
            "as_of_date": as_of_date.isoformat() if as_of_date else "",
        },
        "summary": {
            "total_records": total_positions,
            "findings": len(findings),
            "by_severity": by_severity,
        },
        "findings": findings,
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(document, fh, indent=2, ensure_ascii=False)
    print(f"  JSON -> {path}")
    return path


def write_csv(issues: List[Dict[str, Any]], country: str = "CA") -> str:
    _ensure_output_dir()
    path = os.path.join(OUTPUT_DIR, f"position_integrity_{country}_{_datestamp()}.csv")
    df = _df_from_issues(issues)
    df.to_csv(path, index=False)
    print(f"  CSV  -> {path}")
    return path


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def write_excel(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str = "CA",
    tenant_url: str = "",
    instance_id: str = "",
    as_of_date: datetime.date | None = None,
) -> str:
    _ensure_output_dir()
    path = os.path.join(OUTPUT_DIR, f"position_integrity_{country}_{_datestamp()}.xlsx")

    wb = openpyxl.Workbook()
    normalised = _normalise_dates(issues)

    # ---- Summary sheet (first tab) ----------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(
        ws_sum,
        normalised,
        total_positions,
        country,
        tenant_url,
        instance_id,
        as_of_date=as_of_date,
    )

    # ---- Issues sheet -------------------------------------------------------
    ws = wb.create_sheet(title="Issues")
    _build_issues_sheet(ws, normalised)

    wb.save(path)
    print(f"  XLSX -> {path}")
    return path


def _header_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_issues_sheet(ws, issues: List[Dict[str, Any]]) -> None:
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        _header_cell(ws, 1, col_idx, col_name)

    for row_idx, issue in enumerate(issues, start=2):
        severity = issue.get("Severity", "")
        row_fill = CRITICAL_FILL if severity == "CRITICAL" else HIGH_FILL
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=issue.get(col_name, ""))
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row_idx in range(2, len(issues) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = "A2"


def _build_summary_sheet(
    ws,
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str,
    tenant_url: str = "",
    instance_id: str = "",
    as_of_date: datetime.date | None = None,
) -> None:
    run_date = (as_of_date or datetime.date.today()).isoformat()
    critical_count = sum(1 for i in issues if i.get("Severity") == "CRITICAL")
    high_count = sum(1 for i in issues if i.get("Severity") == "HIGH")
    check_counts = Counter(i.get("Check ID") for i in issues)

    # --- Row 1: Branding header spanning A:D ---
    ws.merge_cells("A1:D1")
    brand_cell = ws["A1"]
    brand_cell.value = f"SF Position Integrity Checker  |  {GITHUB_URL}"
    brand_cell.font = Font(bold=True, color="FFFFFF", size=11)
    brand_cell.fill = HEADER_FILL
    brand_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # --- Row 2: blank spacer ---

    # --- Row 3: original title ---
    ws.merge_cells("A3:D3")
    title_cell = ws["A3"]
    title_cell.value = "SF Position Integrity Checker - Run Summary"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center")

    def label_value(row, label, value):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = SUMMARY_LABEL_FILL
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=row, column=2, value=value)
        vc.alignment = Alignment(horizontal="left")

    instance = _instance_name(tenant_url, instance_id)
    label_value(5, "SF Instance:", instance if instance else " - ")
    label_value(6, "Country:", country)
    label_value(7, "As-of Date:", run_date)
    label_value(8, "Positions Checked:", total_positions)
    label_value(9, "Total Issues Found:", len(issues))
    label_value(10, "CRITICAL Issues:", critical_count)
    label_value(11, "HIGH Issues:", high_count)

    # Check breakdown table header (shifted down by 1 to accommodate instance row)
    for col_idx, label in enumerate(
        ["Check ID", "Description", "Severity", "Count"], start=1
    ):
        cell = ws.cell(row=13, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL

    from validators import CHECK_META

    for r_offset, (chk_id, meta) in enumerate(sorted(CHECK_META.items()), start=1):
        cnt = check_counts.get(chk_id, 0)
        row = 13 + r_offset
        ws.cell(row=row, column=1, value=chk_id)
        ws.cell(row=row, column=2, value=meta.get("description", ""))
        ws.cell(row=row, column=3, value=meta["severity"])
        ws.cell(row=row, column=4, value=cnt)
        row_fill = CRITICAL_FILL if meta["severity"] == "CRITICAL" else HIGH_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = row_fill

    for col, width in zip("ABCD", [12, 45, 12, 8]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------


def write_html(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str = "CA",
    tenant_url: str = "",
    instance_id: str = "",
    as_of_date: datetime.date | None = None,
) -> str:
    _ensure_output_dir()
    path = os.path.join(OUTPUT_DIR, f"position_integrity_{country}_{_datestamp()}.html")
    html = _build_html(
        issues, total_positions, country, tenant_url, instance_id, as_of_date=as_of_date
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML -> {path}")
    return path


def _build_html(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str,
    tenant_url: str = "",
    instance_id: str = "",
    as_of_date: datetime.date | None = None,
) -> str:
    run_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_date = (as_of_date or datetime.date.today()).isoformat()
    instance = _instance_name(tenant_url, instance_id)
    critical_count = sum(1 for i in issues if i.get("Severity") == "CRITICAL")
    high_count = sum(1 for i in issues if i.get("Severity") == "HIGH")

    severities = sorted({i.get("Severity", "") for i in issues} - {""})
    check_ids = sorted({i.get("Check ID", "") for i in issues} - {""})
    categories = sorted({i.get("Check Category", "") for i in issues} - {""})
    employee_statuses = sorted({i.get("Employee Status", "") for i in issues} - {""})
    has_vacant_rows = any(not str(i.get("Employee Status", "")).strip() for i in issues)

    def opts(values):
        return "\n".join(f'<option value="{v}">{v}</option>' for v in values)

    issues = _normalise_dates(issues)

    rows_html_parts = []
    for issue in issues:
        sev = issue.get("Severity", "")
        css = "critical" if sev == "CRITICAL" else "high"
        cells = "".join(f"<td>{issue.get(c, '')}</td>" for c in COLUMNS)
        rows_html_parts.append(f'<tr class="{css}">{cells}</tr>')
    rows_html = "\n".join(rows_html_parts)

    col_headers = "".join(
        f'<th onclick="sortTable({i})">{c} <span class="sort-icon">⇅</span></th>'
        for i, c in enumerate(COLUMNS)
    )

    # Count unique impacted positions by employee status to show a quick fix scope.
    seen_positions = set()
    active_positions = 0
    terminated_positions = 0
    vacant_positions = 0
    other_positions = 0
    for idx, issue in enumerate(issues):
        position_id = str(issue.get("Position ID", "")).strip() or f"__row_{idx}"
        if position_id in seen_positions:
            continue
        seen_positions.add(position_id)

        emp_status = str(issue.get("Employee Status", "")).strip().lower()
        if not emp_status:
            vacant_positions += 1
        elif emp_status == "active":
            active_positions += 1
        elif emp_status == "terminated":
            terminated_positions += 1
        else:
            other_positions += 1

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SF Position Integrity Report - {country} - {run_dt[:10]}</title>
<script src="https://cdn.sheetjs.com/xlsx-latest/package/dist/xlsx.full.min.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
          background: #f4f6f9; color: #333; font-size: 13px; }}
  h1 {{ font-size: 1.4rem; color: #1f3864; margin-bottom: 0.25rem; }}
  .header-bar {{ background: #1f3864; color: #fff; padding: 1rem 1.5rem;
                 display: flex; align-items: center; justify-content: space-between; gap: 1rem; }}
  .header-bar h1 {{ color: #fff; }}
  .header-bar .sub {{ font-size: 0.85rem; opacity: 0.8; }}
  .instance-badge {{ background: rgba(255,255,255,.15); border: 1px solid rgba(255,255,255,.3);
                     border-radius: 6px; padding: 0.35rem 0.75rem; font-size: 0.82rem;
                     font-family: "SFMono-Regular", Consolas, monospace; white-space: nowrap;
                     display: flex; flex-direction: column; align-items: flex-end; gap: 0.1rem; }}
  .instance-badge .lbl {{ font-size: 0.7rem; opacity: 0.7; text-transform: uppercase; letter-spacing: 0.06em; }}
  .instance-badge .val {{ font-weight: 700; font-size: 0.9rem; letter-spacing: 0.02em; }}
  .cards {{ display: flex; gap: 1rem; padding: 1rem 1.5rem; flex-wrap: wrap; }}
  .card {{ background: #fff; border-radius: 8px; padding: 1rem 1.5rem;
           flex: 1; min-width: 150px; box-shadow: 0 1px 4px rgba(0,0,0,.1); text-align: center; }}
  .card .num {{ font-size: 2rem; font-weight: 700; line-height: 1; }}
  .card .lbl {{ font-size: 0.8rem; color: #666; margin-top: 0.3rem; }}
  .card.critical .num {{ color: #c0392b; }}
  .card.high .num {{ color: #d35400; }}
  .card.total .num {{ color: #1f3864; }}
  .card.status-active .num {{ color: #1d7d3a; }}
  .card.status-terminated .num {{ color: #8e3b46; }}
  .card.status-vacant .num {{ color: #6b7280; }}
  .card.status-other .num {{ color: #5b4b8a; }}
  .filters {{ padding: 0 1.5rem 1rem; display: flex; gap: 0.75rem; flex-wrap: wrap;
              align-items: flex-end; }}
  .filters label {{ font-size: 0.8rem; color: #555; display: flex; flex-direction: column;
                    gap: 0.2rem; }}
  .filters select, .filters input {{ padding: 0.35rem 0.6rem; border: 1px solid #ccc;
                                      border-radius: 4px; font-size: 0.85rem; }}
  .filters button {{ padding: 0.38rem 1rem; background: #1f3864; color: #fff;
                     border: none; border-radius: 4px; cursor: pointer; font-size: 0.85rem; }}
  .filters button:hover {{ background: #2e5090; }}
  .export-bar {{ padding: 0 1.5rem 0.75rem; display: flex; gap: 0.5rem; flex-wrap: wrap; }}
  .export-bar span {{ font-size: 0.8rem; color: #555; align-self: center; margin-right: 0.25rem; }}
  .btn-export {{ padding: 0.38rem 0.9rem; border: none; border-radius: 4px; cursor: pointer;
                 font-size: 0.82rem; font-weight: 600; display: inline-flex; align-items: center; gap: 0.35rem; }}
  .btn-csv   {{ background: #217346; color: #fff; }}
  .btn-csv:hover   {{ background: #1a5c38; }}
  .btn-xlsx  {{ background: #1f7244; color: #fff; }}
  .btn-xlsx:hover  {{ background: #175a36; }}
  .btn-pdf   {{ background: #c0392b; color: #fff; }}
  .btn-pdf:hover   {{ background: #96281b; }}
  @media print {{
    .filters, .export-bar, .header-bar .sub, footer {{ display: none !important; }}
    body {{ background: #fff; font-size: 11px; }}
    .cards {{ flex-wrap: nowrap; }}
    .table-wrap {{ overflow: visible; }}
    td, th {{ white-space: normal !important; max-width: none !important; }}
  }}
  .table-wrap {{ padding: 0 1.5rem 2rem; overflow-x: auto; }}
  table {{ border-collapse: collapse; width: 100%; background: #fff;
           box-shadow: 0 1px 4px rgba(0,0,0,.1); border-radius: 6px; overflow: hidden; }}
  thead tr {{ position: sticky; top: 0; z-index: 10; }}
  th {{ background: #1f3864; color: #fff; padding: 0.6rem 0.75rem; text-align: left;
        white-space: nowrap; cursor: pointer; user-select: none; font-size: 0.8rem; }}
  th:hover {{ background: #2e5090; }}
  .sort-icon {{ opacity: 0.6; font-size: 0.7rem; }}
  td {{ padding: 0.45rem 0.75rem; border-bottom: 1px solid #eee; white-space: nowrap;
        max-width: 280px; overflow: hidden; text-overflow: ellipsis; }}
  tr.critical td {{ background: #ffd7d7; }}
  tr.high td {{ background: #ffe4c4; }}
  tr.hidden {{ display: none; }}
  tr:hover td {{ filter: brightness(0.95); }}
  .badge {{ display: inline-block; padding: 0.15rem 0.55rem; border-radius: 999px;
             font-size: 0.72rem; font-weight: 700; letter-spacing: 0.03em; }}
  .badge-critical {{ background: #c0392b; color: #fff; }}
  .badge-high {{ background: #d35400; color: #fff; }}
  footer {{ text-align: center; padding: 1rem; color: #888; font-size: 0.78rem; }}
</style>
</head>
<body>

<div class="header-bar">
  <div>
    <h1>SF Position Integrity Checker</h1>
    <div class="sub">Country: {country} &nbsp;|&nbsp; As-of Date: {run_date} &nbsp;|&nbsp; Generated: {run_dt}</div>
  </div>
  {f'''<div class="instance-badge"><span class="lbl">SF Instance</span><span class="val">{instance}</span></div>''' if instance else ""}
</div>

<div class="cards">
  <div class="card total">
    <div class="num">{total_positions}</div>
    <div class="lbl">Positions Checked</div>
  </div>
  <div class="card total">
    <div class="num">{len(issues)}</div>
    <div class="lbl">Total Issues</div>
  </div>
  <div class="card critical">
    <div class="num">{critical_count}</div>
    <div class="lbl">Critical</div>
  </div>
  <div class="card high">
    <div class="num">{high_count}</div>
    <div class="lbl">High</div>
  </div>
</div>

<div class="cards">
  <div class="card status-active">
    <div class="num">{active_positions}</div>
    <div class="lbl">Active Positions (in findings)</div>
  </div>
  <div class="card status-terminated">
    <div class="num">{terminated_positions}</div>
    <div class="lbl">Terminated Positions (in findings)</div>
  </div>
  <div class="card status-vacant">
    <div class="num">{vacant_positions}</div>
    <div class="lbl">Vacant Positions (in findings)</div>
  </div>
  {f'''<div class="card status-other"><div class="num">{other_positions}</div><div class="lbl">Other Employee Status</div></div>''' if other_positions else ""}
</div>

<div class="filters">
  <label>Severity
    <select id="f-severity" onchange="applyFilters()">
      <option value="">All</option>
      {opts(severities)}
    </select>
  </label>
  <label>Check ID
    <select id="f-checkid" onchange="applyFilters()">
      <option value="">All</option>
      {opts(check_ids)}
    </select>
  </label>
  <label>Check Category
    <select id="f-category" onchange="applyFilters()">
      <option value="">All</option>
      {opts(categories)}
    </select>
  </label>
  <label>Employee Status
    <select id="f-empstatus" onchange="applyFilters()">
      <option value="">All</option>
      {opts(employee_statuses)}
      {'<option value="__vacant__">Vacant</option>' if has_vacant_rows else ""}
    </select>
  </label>
  <label>Search
    <input type="text" id="f-search" placeholder="Search all columns..."
           oninput="applyFilters()">
  </label>
  <button onclick="clearFilters()">Clear</button>
</div>

<div class="export-bar">
  <span>Export:</span>
  <button class="btn-export btn-csv"  onclick="exportCSV()">&#x2B07; CSV</button>
  <button class="btn-export btn-xlsx" onclick="exportExcel()">&#x2B07; Excel</button>
  <button class="btn-export btn-pdf"  onclick="window.print()">&#x1F5A8; Print / PDF</button>
</div>

<div class="table-wrap">
  <table id="issues-table">
    <thead><tr>{col_headers}</tr></thead>
    <tbody id="table-body">
{rows_html}
    </tbody>
  </table>
</div>

<footer>Generated by SF Position Integrity Checker v{VERSION} &nbsp;|&nbsp; As-of Date: {run_date} &nbsp;|&nbsp; Generated: {run_dt} &nbsp;|&nbsp; Country: {country}{f" &nbsp;|&nbsp; Instance: {instance}" if instance else ""}</footer>

<script>
const COL_SEVERITY = {COLUMNS.index("Severity")};
const COL_CHECKID  = {COLUMNS.index("Check ID")};
const COL_CATEGORY = {COLUMNS.index("Check Category")};
const COL_EMPSTATUS = {COLUMNS.index("Employee Status")};

const REPORT_FILENAME = "position_integrity_{country}_{run_dt[:10]}";

let sortDir = {{}};

function visibleRows() {{
  return Array.from(document.querySelectorAll("#table-body tr"))
              .filter(r => !r.classList.contains("hidden"));
}}

function tableData() {{
  const headers = Array.from(document.querySelectorAll("#issues-table thead th"))
                       .map(th => th.textContent.replace(/[⇅↑↓]/g,"").trim());
  const rows = visibleRows().map(tr =>
    Array.from(tr.cells).map(td => td.textContent.trim())
  );
  return {{ headers, rows }};
}}

function exportCSV() {{
  const {{ headers, rows }} = tableData();
  const escape = v => `"${{v.replace(/"/g,'""')}}"`;
  const lines = [headers.map(escape).join(","),
                 ...rows.map(r => r.map(escape).join(","))];
  const blob = new Blob([lines.join("\\r\\n")], {{ type: "text/csv;charset=utf-8;" }});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = REPORT_FILENAME + ".csv";
  a.click();
  URL.revokeObjectURL(a.href);
}}

function exportExcel() {{
  if (typeof XLSX === "undefined") {{
    alert("SheetJS library not loaded. Check your internet connection and reload the page.");
    return;
  }}
  const {{ headers, rows }} = tableData();
  const wb = XLSX.utils.book_new();

  const wsData = [headers, ...rows];
  const ws = XLSX.utils.aoa_to_sheet(wsData);

  const colWidths = headers.map((h, i) => {{
    const max = Math.max(h.length, ...rows.map(r => (r[i] || "").length));
    return {{ wch: Math.min(max + 2, 60) }};
  }});
  ws["!cols"] = colWidths;
  ws["!freeze"] = {{ xSplit: 0, ySplit: 1 }};

  XLSX.utils.book_append_sheet(wb, ws, "Issues");

  const critCount = rows.filter(r => r[COL_SEVERITY] === "CRITICAL").length;
  const highCount = rows.filter(r => r[COL_SEVERITY] === "HIGH").length;
  const summaryData = [
    ["SF Position Integrity Checker - Export Summary"],
    [],
    ["Country",        "{country}"],
    ["Report Date",   "{run_dt[:10]}"],
    ["Visible Issues", rows.length],
    ["CRITICAL",       critCount],
    ["HIGH",           highCount],
  ];
  const wsSummary = XLSX.utils.aoa_to_sheet(summaryData);
  wsSummary["!cols"] = [{{ wch: 22 }}, {{ wch: 18 }}];
  XLSX.utils.book_append_sheet(wb, wsSummary, "Summary");

  XLSX.writeFile(wb, REPORT_FILENAME + ".xlsx");
}}

function cellText(row, col) {{
  return (row.cells[col] ? row.cells[col].textContent : "").toLowerCase();
}}

function applyFilters() {{
  const sev  = document.getElementById("f-severity").value.toLowerCase();
  const chk  = document.getElementById("f-checkid").value.toLowerCase();
  const cat  = document.getElementById("f-category").value.toLowerCase();
  const emp  = document.getElementById("f-empstatus").value.toLowerCase();
  const srch = document.getElementById("f-search").value.toLowerCase();
  const rows = document.querySelectorAll("#table-body tr");
  rows.forEach(row => {{
    const s0 = !sev  || cellText(row, COL_SEVERITY) === sev;
    const s1 = !chk  || cellText(row, COL_CHECKID)  === chk;
    const s2 = !cat  || cellText(row, COL_CATEGORY) === cat;
    const rowEmpStatus = cellText(row, COL_EMPSTATUS);
    const s3 = !emp || (emp === "__vacant__" ? rowEmpStatus === "" : rowEmpStatus === emp);
    const s4 = !srch || Array.from(row.cells).some(c =>
                  c.textContent.toLowerCase().includes(srch));
    row.classList.toggle("hidden", !(s0 && s1 && s2 && s3 && s4));
  }});
}}

function clearFilters() {{
  ["f-severity","f-checkid","f-category","f-empstatus"].forEach(id =>
    document.getElementById(id).value = "");
  document.getElementById("f-search").value = "";
  applyFilters();
}}

function sortTable(col) {{
  const tbody = document.getElementById("table-body");
  const rows  = Array.from(tbody.querySelectorAll("tr"));
  sortDir[col] = !sortDir[col];
  const dir = sortDir[col] ? 1 : -1;
  rows.sort((a, b) => {{
    const ta = cellText(a, col);
    const tb = cellText(b, col);
    return ta < tb ? -dir : ta > tb ? dir : 0;
  }});
  rows.forEach(r => tbody.appendChild(r));
}}
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Console summary table
# ---------------------------------------------------------------------------


def print_console_summary(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str,
    as_of_date: datetime.date | None = None,
) -> None:
    check_counts = Counter(i.get("Check ID") for i in issues)
    critical_n = sum(1 for i in issues if i.get("Severity") == "CRITICAL")
    high_n = sum(1 for i in issues if i.get("Severity") == "HIGH")

    from validators import CHECK_META

    print("\n")
    print(
        "╔══════════╦══════════════════════════════════════════════╦══════════╦═══════════╗"
    )
    print(
        "║ Check ID ║ Description                                  ║ Severity ║   Count   ║"
    )
    print(
        "╠══════════╬══════════════════════════════════════════════╬══════════╬═══════════╣"
    )
    for chk_id in sorted(CHECK_META.keys()):
        cnt = check_counts.get(chk_id, 0)
        sev = CHECK_META[chk_id]["severity"]
        desc = CHECK_META[chk_id].get("description", "")[:44].ljust(44)
        chk = chk_id.ljust(8)
        sev_s = sev.ljust(8)
        cnt_s = str(cnt).center(9)
        print(f"║ {chk} ║ {desc} ║ {sev_s} ║ {cnt_s} ║")
    print(
        "╠══════════╬══════════════════════════════════════════════╬══════════╬═══════════╣"
    )
    print(
        f"║ {'TOTAL'.ljust(8)} ║ {''.ljust(44)} ║ {''.ljust(8)} ║ {str(len(issues)).center(9)} ║"
    )
    print(
        "╚══════════╩══════════════════════════════════════════════╩══════════╩═══════════╝"
    )
    print(f"\n  Positions Checked : {total_positions}")
    print(f"  Total Issues      : {len(issues)}")
    print(f"  CRITICAL          : {critical_n}")
    print(f"  HIGH              : {high_n}")
    print(f"  Country           : {country}")
    print(f"  As-of Date        : {(as_of_date or datetime.date.today()).isoformat()}")



# ---------------------------------------------------------------------------
# Fix pack
# ---------------------------------------------------------------------------

ISSUE_OWNERS = {
    "foundation": "Foundation data owner",
    "job": "Job architecture owner",
    "cost": "Finance cost centre owner",
    "location": "Workplace location owner",
    "legal": "HR legal entity owner",
    "employee": "HR operations owner",
}


def _owner_for_issue(issue: Dict[str, Any]) -> str:
    text = " ".join(
        str(issue.get(k, "")).lower()
        for k in ("Check Category", "Failed Field", "Issue Description", "Check ID")
    )
    for key, owner in ISSUE_OWNERS.items():
        if key in text:
            return owner
    return "Position data owner"


def _fix_action(issue: Dict[str, Any]) -> str:
    field = issue.get("Failed Field") or "Position field"
    check_id = issue.get("Check ID") or "rule"
    return f"Review {field} for {check_id} and update the source position record or related foundation object."


def write_fix_pack(issues: List[Dict[str, Any]], country: str = "CA") -> str:
    """Write suggested correction templates and ownership by issue type."""
    _ensure_output_dir()
    path = os.path.join(OUTPUT_DIR, f"position_integrity_fix_pack_{country}_{_datestamp()}.csv")
    rows = []
    for issue in _normalise_dates(issues):
        rows.append(
            {
                "Owner": _owner_for_issue(issue),
                "Issue Type": issue.get("Check ID", ""),
                "Severity": issue.get("Severity", ""),
                "Position ID": issue.get("Position ID", ""),
                "Effective Start Date": issue.get("Effective Start Date", ""),
                "Failed Field": issue.get("Failed Field", ""),
                "Current Value": "",
                "Suggested Correction": "",
                "Correction Template": _fix_action(issue),
                "Notes": issue.get("Issue Description", ""),
            }
        )
    df = pd.DataFrame(rows, columns=[
        "Owner",
        "Issue Type",
        "Severity",
        "Position ID",
        "Effective Start Date",
        "Failed Field",
        "Current Value",
        "Suggested Correction",
        "Correction Template",
        "Notes",
    ])
    df.to_csv(path, index=False)
    print(f"  FIX  -> {path}")
    return path

# ---------------------------------------------------------------------------
# Run manifest
# ---------------------------------------------------------------------------


def write_run_manifest(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str,
    tenant_url: str = "",
    checks_disabled: List[str] = None,
    as_of_date: datetime.date | None = None,
) -> str:
    """Write output/run_manifest.json after every validation run."""
    _ensure_output_dir()
    path = os.path.join(OUTPUT_DIR, "run_manifest.json")

    from validators import _ENABLED_RULES, _ALL_RULES

    checks_run = sorted(r["id"] for r in _ENABLED_RULES)
    all_check_ids = {r["id"] for r in _ALL_RULES}
    enabled_ids = set(checks_run)
    checks_disabled = checks_disabled or sorted(all_check_ids - enabled_ids)

    visible = _visible_issues(issues)
    hidden_n = len(issues) - len(visible)
    critical_n = sum(1 for i in issues if i.get("Severity") == "CRITICAL")
    high_n = sum(1 for i in issues if i.get("Severity") == "HIGH")

    manifest = {
        "tool": "SF Position Integrity Checker",
        "version": VERSION,
        "run_timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
        "tenant_url": _mask_tenant_url(tenant_url) if tenant_url else "",
        "country": country,
        "as_of_date": (as_of_date or datetime.date.today()).isoformat(),
        "positions_checked": total_positions,
        "total_issues": len(issues),
        "hidden_issues_count": hidden_n,
        "critical_issues": critical_n,
        "high_issues": high_n,
        "checks_run": checks_run,
        "checks_disabled": checks_disabled,
        "fix_pack": f"position_integrity_fix_pack_{country}_{_datestamp()}.csv",
        "ownership_model": "Owner is derived from issue category, failed field and rule text.",
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    print(f"  JSON -> {path}")
    return path


# ---------------------------------------------------------------------------
# Master write function
# ---------------------------------------------------------------------------


def write_report_meta(country: str, instance_id: str) -> None:
    """Write a per-report sidecar <stem>.meta.json so the web UI can show
    the correct instance name for historical reports regardless of what
    instance is currently configured."""
    _ensure_output_dir()
    stem = f"position_integrity_{country}_{_datestamp()}"
    path = os.path.join(OUTPUT_DIR, f"{stem}.meta.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"instance_id": instance_id}, f)


def write_all_reports(
    issues: List[Dict[str, Any]],
    total_positions: int,
    country: str = "CA",
    tenant_url: str = "",
    instance_id: str = "",
    as_of_date: datetime.date | None = None,
) -> None:
    print(f"\n[REPORT] Writing output files to ./{OUTPUT_DIR}/")
    visible = _visible_issues(issues)
    write_csv(visible, country)
    write_excel(
        visible,
        total_positions,
        country,
        tenant_url=tenant_url,
        instance_id=instance_id,
        as_of_date=as_of_date,
    )
    write_html(
        visible,
        total_positions,
        country,
        tenant_url=tenant_url,
        instance_id=instance_id,
        as_of_date=as_of_date,
    )
    write_fix_pack(visible, country)
    write_findings_json(
        visible,
        total_positions,
        country,
        tenant_url=tenant_url,
        as_of_date=as_of_date,
    )
    write_run_manifest(
        issues, total_positions, country, tenant_url=tenant_url, as_of_date=as_of_date
    )
    write_report_meta(country, instance_id)
    print_console_summary(issues, total_positions, country, as_of_date=as_of_date)
